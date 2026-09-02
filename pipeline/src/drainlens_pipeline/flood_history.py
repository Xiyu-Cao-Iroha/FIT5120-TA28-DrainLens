"""Recorded flood incidents by named area, for the historical flood board.

Epic 2 asks which areas have the highest recorded flood-related incident
counts. Two published files answer that and neither can do it alone, which is
the whole reason this module exists rather than a fetch in the browser.

**What a count is.** One Victoria SES crew dispatch. The Data Quality
Statement is explicit: *"Tasks may or may not have been subsequently undertaken
in relation to each dispatch"*, and *"the figures preclude multiple crew
attendances at any one incident"*. It is not a count of floods, of damaged
properties, or of calls. AC 2.3.1.c turns on saying so.

**Why Flood alone.** The DQS puts *"heavy rain leading to flash flooding"*
inside the **Storm** definition, so the incident type closest to urban
stormwater is split across two buckets and the larger one also holds wind,
hail, dust and snow. Ranking on Storm, or on Storm plus Flood, produces the
Dandenongs and the Yarra Valley — where the dispatches are for fallen limbs.
Greater Melbourne's top eight by each measure share two names. So the board
ranks on Flood and the page says what that leaves out; the alternative is a
wind-damage map with a flood heading.

**Why Greater Melbourne.** Not one Melbourne area reaches the statewide top
five, which are regional river towns, while Greater Melbourne holds 45.1% of
the flood incidents. A board for Melbourne residents that opens with Mildura
has answered a question its reader did not ask. The scope is on the artefact
and belongs on the page.

**Why SA2.** The DQS says in as many words that "SA1 regions are not named",
and AC 2.1.1.d requires an area name. The 7-digit code carries the SA2
identifier inside it, and ABS ASGS 2011 turns that into a name — see
`join`, where the match is asserted to be total rather than assumed.

**Why every year is kept.** 2010-11 alone is 41.8% of Greater Melbourne's
six-year total, and dropping it changes two of the top five. The ranking uses
the six-year total, because that is what a recorded count is and choosing a
window would be choosing the answer; the per-year series ships beside it so a
reader can see that one summer rather than being told about it in a footnote.

Details and the measurements behind all of this: `docs/FLOOD-HISTORY-DATA.md`.
"""

from __future__ import annotations

import csv
import io
import json
import urllib.request
import zipfile
from dataclasses import dataclass
from typing import Callable, Iterable, Mapping, Sequence

#: The incident counts.
SOURCE = {
    "dataset": "VICSES Incidents Per SA1 ABS Census Areas, 2009 - 2015",
    "publisher": "Victoria State Emergency Service",
    "licence": "CC BY 4.0",
    "dataset_id": "victoria-ses-incidents-per-sa1-abs-census-areas-2009-2015",
}

#: The names, which the incident file does not have.
GEOGRAPHY = {
    "dataset": "Australian Statistical Geography Standard (ASGS) 2011, Volume 1",
    "publisher": "Australian Bureau of Statistics",
    "licence": "CC BY 2.5 AU",
    "dataset_id": "1270.0.55.001",
}

#: Where each file actually lives.
#:
#: data.vic's catalogue record for the incidents points at a URL that returns a
#: 404 HTML page, which `curl` reports as a successful 119 KB download. The
#: live copy is a ZIP on the SES site holding the workbook and its Data Quality
#: Statement. `fetch` checks the content rather than the status code.
INCIDENTS_URL = (
    "https://www.ses.vic.gov.au/documents/d/www/incidents-per-sa1-abs-census-areas?download=true"
)
GEOGRAPHY_URL = (
    "https://www.abs.gov.au/ausstats/subscriber.nsf/log?openagent"
    "&1270055001_sa1_2011_aust_csv.zip&1270.0.55.001&Data%20Cubes"
    "&5AD36D669F284E70CA257801000C69BE&0&July%202011&23.12.2010&Latest"
)

#: The six financial years the workbook carries, in order.
#:
#: The DQS gives the reference period as 1 July 2009 to 30 June 2015 and all
#: six are complete. The catalogue's "July 2009 - 8 August 2015" is the
#: collection window; the extra five weeks fall in 2015-16 and are in no
#: column. Asserted against the file, so a republished workbook with a
#: different span fails here rather than shifting the meaning of a total.
YEARS: tuple[str, ...] = ("2009-10", "2010-11", "2011-12", "2012-13", "2013-14", "2014-15")

REPORTING_PERIOD = {"start": "2009-07-01", "end": "2015-06-30", "years": list(YEARS)}

#: The incident type ranked on, and the type its definition excludes.
INCIDENT_TYPE = "Flood"
FLASH_FLOODING_TYPE = "Storm"

#: The scope, as ABS names it in `GCCSA_NAME_2011`.
SCOPE = "Greater Melbourne"

#: AC 2.1.1.b, and AC 2.2.1.b's ceiling. Nothing beyond 30 is ever displayed,
#: so nothing beyond 30 is published.
DEFAULT_AREAS = 5
MAX_AREAS = 30

#: How the workbook marks a count withheld under the Privacy and Data
#: Protection Act 2014 — regions of 20 people or fewer, or 10 dwellings or
#: fewer, where a count could identify somebody.
SUPPRESSED = "X"

#: The SA2 containing the pilot extent, so the board can offer the map without
#: the offer being abstract.
PILOT_AREA = "Kensington"


class FloodHistoryError(RuntimeError):
    """The sources do not support the claim the artefact would make."""


@dataclass(frozen=True)
class Place:
    """Where an SA1 sits: the named area, and the capital-city region."""

    area: str
    greater_capital: str


@dataclass(frozen=True)
class Region:
    """One SA1: its code, its yearly counts, and whether any were withheld."""

    code: str
    by_year: tuple[int, ...]
    suppressed: bool


@dataclass(frozen=True)
class Area:
    """One named SA2, aggregated from the regions inside it."""

    name: str
    by_year: tuple[int, ...]
    regions: int
    suppressed_regions: int

    @property
    def total(self) -> int:
        return sum(self.by_year)

    @property
    def complete(self) -> bool:
        """False where a count inside this area was withheld for privacy.

        The total is then a lower bound, and AC 2.1.1.g and AC 2.2.1.f require
        the page to say so. Aggregating to SA2 does not resolve a suppressed
        SA1; it conceals one, which is why the flag travels with the area
        rather than being dropped at the join.
        """
        return self.suppressed_regions == 0


def read_incidents(workbook: bytes) -> list[Region]:
    """Every SA1's flood counts, read from the published workbook.

    The sheet has two header rows: the incident type spread across its six
    year columns, and the years themselves. Both are checked, because a
    republished file that moved a column would otherwise be read as a
    different incident type without anything failing.
    """
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(workbook), read_only=True, data_only=True)
    try:
        if "Master" not in book.sheetnames:
            raise FloodHistoryError(
                f"expected a 'Master' sheet, found {book.sheetnames}"
            )
        rows = list(book["Master"].iter_rows(values_only=True))
    finally:
        book.close()

    if len(rows) < 5:
        raise FloodHistoryError(f"the Master sheet holds {len(rows)} rows, too few to be the data")

    types, years = rows[2], rows[3]
    if str(years[0]).strip() != "region_id":
        raise FloodHistoryError(f"expected 'region_id' in the first column, found {years[0]!r}")

    columns = [
        i
        for i, (kind, year) in enumerate(zip(types, years))
        if kind == INCIDENT_TYPE and year in YEARS
    ]
    found = tuple(str(years[i]) for i in columns)
    if found != YEARS:
        raise FloodHistoryError(
            f"expected the {INCIDENT_TYPE} columns to be {YEARS}, found {found}"
        )

    regions: list[Region] = []
    for row in rows[4:]:
        if row[0] is None:
            continue
        code = str(row[0]).strip()
        if len(code) != 7 or not code.isdigit():
            raise FloodHistoryError(
                f"region_id {code!r} is not a 7-digit SA1 code; the ABS join reads it as "
                "State, SA2 and SA1 and would silently mis-join"
            )
        counts, withheld = [], False
        for i in columns:
            value = row[i]
            if isinstance(value, str):
                if value.strip() != SUPPRESSED:
                    raise FloodHistoryError(
                        f"region {code} holds {value!r}, which is neither a count nor "
                        f"the documented suppression marker {SUPPRESSED!r}"
                    )
                withheld = True
                counts.append(0)
            else:
                counts.append(int(value or 0))
        regions.append(Region(code, tuple(counts), withheld))

    if not regions:
        raise FloodHistoryError("the workbook carried no regions")
    codes = {r.code for r in regions}
    if len(codes) != len(regions):
        raise FloodHistoryError(f"{len(regions) - len(codes)} region_ids appear more than once")
    return regions


def read_geography(allocation: bytes, *, state: str = "Victoria") -> dict[str, Place]:
    """SA1 7-digit code to its SA2 name and greater-capital-city area.

    The whole state, not the scope. Scope is a filter applied after the join,
    because the two questions are different: *does this region resolve to a
    real SA1* is an integrity check on the pair of files, and *is it in Greater
    Melbourne* is an editorial decision. Filtering first collapses them, and
    then a genuine mis-join looks exactly like a region in the country.
    """
    places: dict[str, Place] = {}
    text = allocation.decode("utf-8-sig")
    for record in csv.DictReader(io.StringIO(text)):
        if record["STATE_NAME_2011"] != state:
            continue
        places[record["SA1_7DIGITCODE_2011"]] = Place(
            record["SA2_NAME_2011"], record["GCCSA_NAME_2011"]
        )
    if not places:
        raise FloodHistoryError(
            f"no SA1 in the ABS allocation belongs to {state!r}; the name must match "
            "STATE_NAME_2011 exactly"
        )
    return places


def join(
    regions: Sequence[Region],
    places: Mapping[str, Place],
    *,
    scope: str = SCOPE,
) -> list[Area]:
    """Aggregate the regions to named areas inside the scope.

    The check that matters happens first and applies to **every** region, in
    scope or not: each one must resolve to a real SA1 in the allocation. The
    Data Quality Statement describes the 7-digit code's structure, but a
    description is not evidence; a total join is. Measured on the published
    files, all 13,339 regions match and no Victorian SA1 is left over, which is
    what makes an SA2 name safe to print beside a count.

    Only then is the scope applied. A region outside it is dropped knowingly.
    """
    unmatched = [r.code for r in regions if r.code not in places]
    if unmatched:
        raise FloodHistoryError(
            f"{len(unmatched)} of {len(regions)} regions have no SA2 name in the ABS "
            f"allocation (first: {unmatched[:3]}). Ranking would silently drop them, "
            "and a missing area is indistinguishable from an area with no incidents "
            "once it is gone"
        )

    in_scope = [r for r in regions if places[r.code].greater_capital == scope]
    if not in_scope:
        raise FloodHistoryError(
            f"no region belongs to {scope!r}; the scope must match GCCSA_NAME_2011 exactly"
        )

    totals: dict[str, list[int]] = {}
    counts: dict[str, list[int]] = {}
    for region in in_scope:
        name = places[region.code].area
        if name not in totals:
            totals[name] = [0] * len(YEARS)
            counts[name] = [0, 0]
        for i, value in enumerate(region.by_year):
            totals[name][i] += value
        counts[name][0] += 1
        counts[name][1] += 1 if region.suppressed else 0

    return [
        Area(name, tuple(totals[name]), counts[name][0], counts[name][1])
        for name in sorted(totals)
    ]


def rank(areas: Iterable[Area], *, limit: int = MAX_AREAS) -> list[Area]:
    """Highest recorded count first, ties broken by name so a build repeats.

    Areas with no recorded flood incident are dropped rather than ranked last.
    A zero here means the SES recorded no flood dispatch in six years, which is
    worth saying about a place somebody asked about and is not worth a row in a
    list of the most affected.
    """
    with_incidents = [a for a in areas if a.total > 0]
    return sorted(with_incidents, key=lambda a: (-a.total, a.name))[:limit]


def build(regions: Sequence[Region], places: Mapping[str, Place]) -> dict:
    """The artefact the board reads."""
    areas = join(regions, places)
    ranked = rank(areas)
    if len(ranked) < DEFAULT_AREAS:
        raise FloodHistoryError(
            f"only {len(ranked)} areas recorded a flood incident; AC 2.1.1.b asks for "
            f"{DEFAULT_AREAS}"
        )

    scope_total = sum(a.total for a in areas)
    pilot = next((a for a in areas if a.name == PILOT_AREA), None)

    # Ties are common and one of them straddles the boundary of the default
    # view: ranks five and six both recorded 133. A reader shown five rows has
    # no way to know a sixth area is level with the last of them, and "the five
    # highest" quietly becomes "five of the six highest". The flag lets the page
    # say so; hiding it would be a ranking that reads as more precise than the
    # counts behind it.
    tied = {
        i
        for i, a in enumerate(ranked)
        if (i > 0 and ranked[i - 1].total == a.total)
        or (i + 1 < len(ranked) and ranked[i + 1].total == a.total)
    }

    return {
        "artefact": "flood-history",
        "version": 1,
        "basis": "sourceProvided",
        "note": (
            "Counts of Victoria SES crew dispatches recorded as Flood, by ABS SA2 area, "
            "over six financial years. One count is one dispatch: a task may or may not "
            "have followed, and several crews at one incident count once. It is not a "
            "measure of severity, damage, or of how much water there was, and it does "
            "not describe current or future conditions."
        ),
        "source": SOURCE,
        "geographySource": GEOGRAPHY,
        "reportingPeriod": REPORTING_PERIOD,
        "geography": {"unit": "SA2", "standard": "ASGS 2011", "scope": SCOPE},
        "incidentType": INCIDENT_TYPE,
        "excludes": (
            f"Flash flooding is recorded under {FLASH_FLOODING_TYPE} rather than "
            f"{INCIDENT_TYPE} and is not counted here."
        ),
        "defaultAreas": DEFAULT_AREAS,
        "counts": {
            "areasPublished": len(ranked),
            "areasWithIncidents": sum(1 for a in areas if a.total > 0),
            "areasInScope": len(areas),
            "regions": len(regions),
            "suppressedRegions": sum(1 for r in regions if r.suppressed),
            "incidentsInScope": scope_total,
            "incidentsPublished": sum(a.total for a in ranked),
        },
        "areas": [
            {
                "rank": i,
                "name": a.name,
                "total": a.total,
                "byYear": list(a.by_year),
                "regions": a.regions,
                "suppressedRegions": a.suppressed_regions,
                "complete": a.complete,
                "tied": i - 1 in tied,
            }
            for i, a in enumerate(ranked, 1)
        ],
        "pilotArea": None
        if pilot is None
        else {
            "name": pilot.name,
            "total": pilot.total,
            "byYear": list(pilot.by_year),
            "complete": pilot.complete,
        },
    }


def _open(url: str, timeout: float) -> bytes:
    with urllib.request.urlopen(url, timeout=timeout) as response:  # noqa: S310
        return response.read()


def fetch(
    *,
    opener: Callable[[str], bytes] | None = None,
    timeout: float = 300.0,
) -> tuple[bytes, bytes]:
    """The two source files, checked for being what they claim to be.

    Both arrive as ZIPs and both publishers have served an HTML error page with
    a 200-shaped response, so the archive is opened here rather than trusted.
    """
    get = opener or (lambda url: _open(url, timeout))

    def only_member(payload: bytes, suffix: str, what: str) -> bytes:
        if not payload.startswith(b"PK"):
            head = payload[:120].decode("utf-8", "replace")
            raise FloodHistoryError(
                f"the {what} download is not a ZIP archive. First bytes: {head!r}"
            )
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            members = [n for n in archive.namelist() if n.lower().endswith(suffix)]
            if not members:
                raise FloodHistoryError(
                    f"the {what} archive holds no {suffix} file: {archive.namelist()}"
                )
            return archive.read(members[0])

    workbook = only_member(get(INCIDENTS_URL), ".xlsx", "incidents")
    allocation = only_member(get(GEOGRAPHY_URL), ".csv", "geography")
    return workbook, allocation


def main(argv: list[str] | None = None) -> int:
    import argparse
    import sys
    from pathlib import Path

    parser = argparse.ArgumentParser(
        prog="python -m drainlens_pipeline.flood_history",
        description="Build the historical flood incident artefact.",
    )
    parser.add_argument("--incidents", type=Path, help="local .xlsx instead of fetching")
    parser.add_argument("--geography", type=Path, help="local ABS .csv instead of fetching")
    parser.add_argument(
        "--out", type=Path, default=Path("../apps/web/public/data/flood-history.json")
    )
    args = parser.parse_args(argv)

    if (args.incidents is None) != (args.geography is None):
        print("give both --incidents and --geography, or neither", file=sys.stderr)
        return 1

    if args.incidents is None:
        print("fetching both sources...")
        workbook, allocation = fetch()
    else:
        workbook = args.incidents.read_bytes()
        allocation = args.geography.read_bytes()

    regions = read_incidents(workbook)
    places = read_geography(allocation)
    artefact = build(regions, places)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(artefact, separators=(",", ":")), encoding="utf-8")

    counts = artefact["counts"]
    print(f"wrote {args.out}  ({args.out.stat().st_size / 1024:.1f} KB)")
    print(f"  SA1 regions read          {counts['regions']:>7,}")
    print(f"  suppressed for privacy    {counts['suppressedRegions']:>7,}")
    print(f"  {SCOPE} areas    {counts['areasInScope']:>7,}")
    print(f"  with a flood incident     {counts['areasWithIncidents']:>7,}")
    print(f"  published (AC 2.2.1.b)    {counts['areasPublished']:>7,}")
    print(f"  incidents in scope        {counts['incidentsInScope']:>7,}")
    for area in artefact["areas"][:DEFAULT_AREAS]:
        flag = "" if area["complete"] else "  (a count inside it was withheld)"
        print(f"    {area['rank']}. {area['name']:<28} {area['total']:>5}{flag}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
